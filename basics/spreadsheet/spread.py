import openpyxl as xl

wb = xl.load_workbook("transactions.xlsx")

sheet = wb['Sheet1']

for row in range(2, sheet.max_row+1):
    t_id_cell = sheet.cell(row, 1)
    pro_id_cell = sheet.cell(row, 2)
    pri_cell = sheet.cell(row, 3)

    if t_id_cell.value is not None:
        Vt_id_cell = t_id_cell.value * 1.3
    else:
        Vt_id_cell = None

    if pro_id_cell.value is not None:
        Vpro_id_cell = pro_id_cell.value * 2
    else:
        Vpro_id_cell = None

    if pri_cell.value is not None:
        Vpri_cell = pri_cell.value * 0.5
    else:
        Vpri_cell = None

    t_id_cell.value = Vt_id_cell
    pro_id_cell.value = Vpro_id_cell
    pri_cell.value = Vpri_cell

wb.save("transactions.xlsx")
